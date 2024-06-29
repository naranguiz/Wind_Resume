#Librerias

import pandas as pd
import openpyxl
import openpyxl.styles

################################README##########################################################

h = 60  #Altura de viento estudiada en [m].

################################################################################################


#Path con nombre del archivo proveniente del explorador eólico
path_viento1 = 'C:\\Users\\Draft Sports\\Desktop\\Análisis2004_2016\\VientoRecon\\recon_{}.csv'.format(h)
path_viento2  = 'C:\\Users\\Draft Sports\\Desktop\\Análisis2004_2016\\VientoRecon\\mean_mes_hora_{}.xlsx'.format(h)

#Lectura CSV proviente del explorador solar
df_viento_recon_h = pd.read_csv(path_viento1, na_values=[" "], header=0, names=['Fecha/Hora', str(h)])

#Interpolación Lineal de datos faltantes

df_viento_recon_h = df_viento_recon_h.interpolate()

#Botar filas que no se utilizarán (1984 a 2006) tal que el análisis será sobre los años 2007-2015.
df_viento_recon_h = df_viento_recon_h.drop(df_viento_recon_h.iloc[0:236684].index, axis=0)

#Conversión de columnas "Fecha/Hora" a formato de fecha y hora.
df_viento_recon_h['Fecha/Hora'] = pd.to_datetime(df_viento_recon_h['Fecha/Hora'])

#Crear una nueva columna con el mes de cada fecha
df_viento_recon_h['Mes'] = df_viento_recon_h['Fecha/Hora'].dt.month


#Calcular el promedio de las columnas altura del viento, por mes y hora.
mean_mes_hora_h = df_viento_recon_h.groupby(['Mes', df_viento_recon_h['Fecha/Hora'].dt.hour])[str(h)].mean()

#Asignamos atributos de df a la lista calculada anteriormente.
df_resultado_h = pd.DataFrame(mean_mes_hora_h)

#Exportar a Excel.
df_resultado_h.to_excel('C:\\Users\\Draft Sports\\Desktop\\Análisis2004_2016\\VientoRecon\\mean_mes_hora_{}.xlsx'.format(h), index=True)

#Preparación Documento Excel

df_prom_mes_hr = pd.read_excel(path_viento2)

# Eliminar las dos primeras columnas
df_prom_mes_hr.drop(df_prom_mes_hr.columns[[0, 1]], axis=1, inplace=True)

# Dividir la tercera columna en intervalos de 24 elementos
df_prom_mes_hr_final = pd.DataFrame([df_prom_mes_hr.iloc[i:i+24, 0].values for i in range(0, len(df_prom_mes_hr), 24)])

# Redondear los números a dos decimales
df_prom_mes_hr_final = df_prom_mes_hr_final.applymap(lambda x: round(x, 2))

# Agregar la columna "Mes" al nuevo DataFrame
meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'] * 24
df_prom_mes_hr_final.insert(0, 'Mes/Hora', meses[:len(df_prom_mes_hr_final)])

# Exportar el nuevo DataFrame a un archivo de Excel
writer = pd.ExcelWriter('promedio_vientorecon_{}[m].xlsx'.format(h), engine='openpyxl')
df_prom_mes_hr_final.to_excel(writer, index=False)


# Estilo de las celdas
workbook = writer.book
worksheet = writer.sheets['Sheet1']
# Diccionario para contar las celdas por color y por fila
color_count = {}

for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=25):
    # Obtener el valor del mes en la primera columna de la fila
    mes = row[0].value
    
    for c in row:
        c.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), 
                                          right=openpyxl.styles.Side(style='thin'), 
                                          top=openpyxl.styles.Side(style='thin'), 
                                          bottom=openpyxl.styles.Side(style='thin'))
        # Agregar fondo de color según los rangos especificados
        value = c.value
        if isinstance(value, float):
            if 6.76 <= value <= 7.25:
                color = '92CDDC'
            elif 7.26 <= value <= 7.75:
                color = '31869B'
            elif 7.76 <= value <= 8.25:
                color = 'FCD5B4'
            elif 8.26 <= value <= 8.75:
                color = 'F79646'
            elif 8.76 <= value <= 9.25:
                color = 'E26B0A'
            elif 9.26 <= value <= 9.75:
                color = 'DD4B4B'
            elif 9.76 <= value <= 10.25:
                color = 'FF0000'
            elif 10.26 <= value <= 10.75:
                color = 'A5251B'
            
            # Contar las celdas por color y por fila
            if mes in color_count:
                if color in color_count[mes]:
                    color_count[mes][color] += 1
                else:
                    color_count[mes][color] = 1
            else:
                color_count[mes] = {color: 1}
            
            # Asignar el fondo de color a la celda
            c.fill = openpyxl.styles.PatternFill(start_color=color, end_color=color, fill_type='solid')

# Obtener el número total de celdas por color y por fila y marcar la celda correspondiente
row_counter = 17
for i, (mes, colors) in enumerate(color_count.items(), start=2):
    worksheet.cell(row=row_counter, column=1, value=mes).fill = openpyxl.styles.PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    for j, (color, count) in enumerate(colors.items(), start=2):  # Cambio en el valor de inicio de la columna
        worksheet.cell(row=row_counter, column=j, value=count).fill = openpyxl.styles.PatternFill(start_color=color, end_color=color, fill_type='solid')
    row_counter += 1

writer.save()
